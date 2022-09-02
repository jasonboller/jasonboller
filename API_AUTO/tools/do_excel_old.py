from openpyxl import load_workbook
from API_AUTO.tools.read_config import ReadConfig
from API_AUTO.tools import project_path
from API_AUTO.tools.get_data import GetData
from API_AUTO.tools.do_regx import DoRegx

class DoExcel:
    @classmethod
    def get_data(cls,file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(project_path.case_config_path, 'MODE', 'mode'))

        tel = getattr(GetData, 'tel_1')# 从get_data里面获取手机号(Excel）

        # 利用python查询数据库的方式，去拿到最大的手机号--这里可以加，也可以放到get_data

        test_data = []
        for key in mode: # 遍历这个存在配置文件里面的字典
            sheet = wb[key]  # 表单名
            if mode[key] == 'all':
                for i in range(2, sheet.max_row+1):
                    sub_data={}
                    sub_data['case_id'] = sheet.cell(i, 1).value
                    sub_data['method'] = sheet.cell(i, 2).value
                    sub_data['url'] = sheet.cell(i, 3).value
                    if sheet.cell(i, 4).value.find('${tel_1}') != -1: # 有找到这个${tel_1}
                        sub_data['data'] = sheet.cell(i, 4).value.replace('${tel_1}', str(tel))
                    else: # 如果没有找到的话
                        sub_data['data'] = DoRegx.do_regx(sheet.cell(i, 4).value)
                    sub_data['title'] = sheet.cell(i, 5).value
                    sub_data['expected'] = sheet.cell(i, 6).value   # 添加一个期望值到测试数据里去
                    sub_data['sheet_name'] = key
                    test_data.append(sub_data)
                    cls.update_tel(tel+2, file_name, 'init') #更新手机号
                    # 这里也是可以优化的？什么时候对手机号去进行更新？更新的手机号是+1，还是+2，+3？
            else:
                for case_id in mode[key]:
                    sub_data = {}
                    sub_data['case_id'] = sheet.cell(case_id+1, 1).value
                    sub_data['method'] = sheet.cell(case_id+1, 2).value
                    sub_data['url'] = sheet.cell(case_id+1, 3).value
                    # 作手机号的替换
                    if sheet.cell(case_id+1, 4).value.find('${tel_1}') != -1: # 有找到这个${tel_1}
                        sub_data['data'] = sheet.cell(case_id+1, 4).value.replace('${tel_1}', str(tel))
                    else: # 如果没有找到的话
                        sub_data['data'] = DoRegx.do_regx(sheet.cell(case_id+1, 4).value)
                    sub_data['title'] = sheet.cell(case_id+1, 5).value
                    sub_data['expected'] = sheet.cell(case_id+1, 6).value  # 添加一个期望值到测试数据里去
                    sub_data['sheet_name'] = key
                    test_data.append(sub_data)
                    cls.update_tel(tel+2, file_name, 'init') #更新手机号

        return test_data
    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult): # 专门写回数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name) # 保存结果

    @classmethod
    def update_tel(cls, tel, file_name, sheet_name):
        '''更新Excel里面手机号的数据'''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel # 赋值操作
        wb.save(file_name)

    def get_tel(self): # Excel获取未注册的手机号
        pass

if __name__ == '__main__':
    res_data = DoExcel().get_data(project_path.test_data_path)
    print(res_data)